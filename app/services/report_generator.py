import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models.audit import Audit, AuditResult
from app.config import settings
import os
import json
from datetime import datetime


class ReportGenerator:
    def generate_report(self, audit: Audit, db: Session) -> str:
        """
        Genera un reporte en Excel de la auditoría
        """
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Auditoría"
        
        # Estilos
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        success_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        error_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
        
        # Título del reporte
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"REPORTE DE AUDITORÍA - {audit.filename}"
        title_cell.font = Font(bold=True, size=14, color="10b981")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Información general
        ws['A2'] = "Fecha de Auditoría:"
        ws['B2'] = audit.created_at.strftime("%d/%m/%Y %H:%M")
        ws['A3'] = "Tasa de Cumplimiento:"
        ws['B3'] = f"{audit.compliance_rate}%"
        ws['A4'] = "Requisitos Cumplidos:"
        ws['B4'] = f"{audit.compliant_items} / {audit.total_items}"
        
        # Espacio
        current_row = 6
        
        # Encabezados de la tabla
        headers = ["ID", "Descripción", "Palabras Clave", "Obligatorio", "Estado", "Archivos Encontrados"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row += 1
        
        # Obtener resultados de la auditoría
        checklist_items = audit.checklist_items
        results = db.query(AuditResult).filter(AuditResult.audit_id == audit.id).all()
        
        # Crear diccionario de resultados por checklist_item_id
        results_dict = {result.checklist_item_id: result for result in results}
        
        # Datos
        for item in checklist_items:
            result = results_dict.get(item.id)
            
            ws.cell(row=current_row, column=1, value=item.item_id)
            ws.cell(row=current_row, column=2, value=item.description)
            ws.cell(row=current_row, column=3, value=item.keywords)
            ws.cell(row=current_row, column=4, value="Sí" if item.is_mandatory else "No")
            
            # Estado
            estado_cell = ws.cell(row=current_row, column=5)
            if result and result.found:
                estado_cell.value = "✓ CUMPLE"
                estado_cell.fill = success_fill
                estado_cell.font = Font(color="059669", bold=True)
            else:
                estado_cell.value = "✗ NO CUMPLE"
                estado_cell.fill = error_fill
                estado_cell.font = Font(color="dc2626", bold=True)
            
            # Archivos encontrados
            archivos_cell = ws.cell(row=current_row, column=6)
            if result and result.matched_files:
                try:
                    matched_files = json.loads(result.matched_files)
                    file_names = [f['name'] for f in matched_files]
                    archivos_cell.value = "\n".join(file_names) if file_names else "Ninguno"
                except:
                    archivos_cell.value = "Error al leer archivos"
            else:
                archivos_cell.value = "Ninguno"
            
            archivos_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            current_row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
        
        # Guardar archivo
        os.makedirs(settings.REPORTS_DIR, exist_ok=True)
        report_filename = f"reporte_auditoria_{audit.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(settings.REPORTS_DIR, report_filename)
        
        wb.save(report_path)
        
        return report_path